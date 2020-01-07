from openpyxl import load_workbook
from string import ascii_uppercase

def setStats(player_stats):
    set_stats = []
    for player in player_stats:
        wins, losses = 0, 0
        for record in player:
            if record != None:
                record = record.split('-')
                wins += int(record[0])
                losses += int(record[1])

        if (wins != 0 or losses != 0):
            percent = (wins / (wins + losses)) * 100
            if (round(percent) == percent):
                percent = ' (' + str(round(percent)) + '%)'

            else:
                percent = ' (~' + str(round(percent)) + '%)'
                
            set_stats.append(str(wins) + '-' + str(losses) + percent)

        else: 
            set_stats.append('/')

    return set_stats

filename = 'number_players.txt'
f = open(filename, 'r')
num_players = int(f.read())
f.close()

read_name = 'bcpr.xlsx'
wb = load_workbook(read_name)
ws = wb['Main Stats']

players = []
for i in range(2, num_players + 2):
    c = ws['A' + str(i)]
    players.append(c.value)

letters = ascii_uppercase[1 : num_players + 1]
records = []
for i in range(2, num_players + 2):
    player_record = []
    for letter in letters:
        c = ws[letter + str(i)]
        player_record.append(c.value)
    
    records.append(player_record)

set_stats = setStats(records)
write_name = 'output.txt'
f = open(write_name, 'w')
for counter, player in enumerate(players):
    f.write(player + '\t' + set_stats[counter] + '\n')
f.close()